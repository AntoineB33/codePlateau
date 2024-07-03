import os
import openpyxl
from openpyxl import Workbook
import win32com.client as win32

def create_or_modify_xlsm(file_path, sheet_name, vba_function_name, vba_code):
    # Ensure the file path is absolute
    file_path = os.path.abspath(file_path)

    # Create the workbook if it doesn't exist
    if not os.path.exists(file_path):
        wb = Workbook()
        wb.save(file_path)

    # Load the workbook with keep_vba=True to preserve macros
    try:
        wb = openpyxl.load_workbook(file_path, keep_vba=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
    
    # Add the sheet if it doesn't exist
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)

    wb.save(file_path)
    wb.close()

    try:
        # Open the file with win32com to modify the VBA project
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False  # Set to True if you want to see the Excel window

        workbook = excel.Workbooks.Open(file_path)

        # Get the VBA project
        vba_project = workbook.VBProject

        # Check if the standard module exists, if not, create it
        module_name = "Module1"
        module_exists = False

        for component in vba_project.VBComponents:
            if component.Name == module_name:
                module_exists = True
                module = component
                break

        if not module_exists:
            module = vba_project.VBComponents.Add(1)  # 1 corresponds to vbext_ct_StdModule
            module.Name = module_name

        # Check if the function exists in the module
        code_lines = module.CodeModule.Lines(1, module.CodeModule.CountOfLines)

        if vba_function_name not in code_lines:
            # Append the VBA code
            module.CodeModule.AddFromString(vba_code)

        # Save and close
        workbook.Close(SaveChanges=True)
        excel.Quit()
    except Exception as e:
        print(f"Error handling VBA project: {e}")
        if 'excel' in locals():
            excel.Quit()

# Define the parameters
file_path = "example2.xlsm"
sheet_name = "MySheet"
vba_function_name = "MyFunction"
vba_code = """
Function MyFunction()
    MsgBox "Hello, VBA!"
End Function
"""

# Execute the function
create_or_modify_xlsm(file_path, sheet_name, vba_function_name, vba_code)
