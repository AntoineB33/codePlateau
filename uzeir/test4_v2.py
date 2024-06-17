import os
import pandas as pd
import numpy as np
import math
from plotly.graph_objects import Figure, Scatter
from tkinter import filedialog, Tk
from scipy.signal import find_peaks, peak_prominences

import openpyxl

from scipy.fft import fft

""""///////////////////Variables globales///////////////////"""
poids_min = float("inf")
debut_time = 0
fin_time = 0
indice_debut = 0
indice_fin = 0
int_time = 0.2
seuil_poids = 4
min_bite_duration = 1  # Minimum bite duration in seconds
min_inactivity = 1
min_peak = 0
excel_all_path = r".\Resultats exp bag_couverts\Resultats exp bag_couverts\Tableau récapitulatif - new algo.xlsx"
excel_titles = [
    "Duree_Totale",
    "Poids_Conso",
    "Action",
    "Duree_activite_Totale",
    "Duree_activite_mean",
    "Duree_activite_max",
    "Duree_activite_min",
    "Proportion_activite",
    "Bouchees",
]
"""/////////////////////////////////////////////////////////"""

root = Tk()
root.withdraw()

dossier = r".\Resultats exp bag_couverts\Resultats exp bag_couverts\28_05_24_xlsx"
# dossier = r".\data_du_bureau\xlsx"

dossier_graphique = r".\Resultats exp bag_couverts\Resultats exp bag_couverts\28_05_24_graph"

date_folder = "_28_05_24"

fichiers = []
for f in os.listdir(dossier):
    if f.endswith(".xlsx") and (f == "2Plateaux-P2-bag.xlsx" or 10):
        fichiers.append(os.path.join(dossier, f))


Tableau_Final = pd.DataFrame(
    columns=[
        "Duree_Totale",
        "Poids_Conso",
        "Action",
        "Duree_activite_Totale",
        "Duree_activite_mean",
        "Duree_activite_max",
        "Duree_activite_min",
        "Proportion_activite_%",
        "Bouchees",
        "Num_fichier",
    ]
)

print(dossier)
new_excel = dict()


# Fonction pour convertir le temps en minutes et secondes
def convert_time(seconds):
    minutes = seconds // 60
    seconds = int(seconds) % 60
    return f"{minutes} min {seconds} s"


def extract_features(segment):
    features = {}
    weights = segment["Ptot"].values
    times = segment["time"].values

    features["Duration"] = times[-1] - times[0]
    features["MaxWeight"] = np.max(weights)
    features["MinWeight"] = np.min(weights)
    features["MeanWeight"] = np.mean(weights)
    features["StdDevWeight"] = np.std(weights)

    # Calculate second derivative
    if len(weights) > 2:
        second_derivative = np.diff(np.diff(weights))
        features["SecondDerivative"] = np.mean(second_derivative)
    else:
        features["SecondDerivative"] = 0

    # Frequency components using FFT
    freq_components = np.abs(fft(weights))
    features["FrequencyComponent1"] = (
        freq_components[1] if len(freq_components) > 1 else 0
    )
    features["FrequencyComponent2"] = (
        freq_components[2] if len(freq_components) > 2 else 0
    )

    # Skewness and Kurtosis
    features["Skewness"] = pd.Series(weights).skew()
    features["Kurtosis"] = pd.Series(weights).kurtosis()

    # Peak analysis
    peaks, properties = find_peaks(weights, prominence=1)
    features["PeakCount"] = len(peaks)
    features["PeakProminence"] = (
        np.mean(properties["prominences"]) if "prominences" in properties else 0
    )

    return features


# Traitement des fichiers
for fichier in fichiers:
    print(fichier)
    df = pd.read_excel(fichier)
    df.columns = ["time", "Ptot"]
    # df = df[df["Ptot"] > 100]
    # df["Ptot"] = np.abs(df["Ptot"])
    # df = df[df["Ptot"] < 3000]
    df["time"] = df["time"] / 1000

    # Filter smoothing time
    indices_time = [df.index[0]]

    for i in range(1, len(df)):
        if df["time"].iloc[i] - df["time"].iloc[indices_time[-1]] >= int_time:
            indices_time.append(i)

    df = df.iloc[indices_time].reset_index(drop=True)

    if df.empty:
        continue  # Skip further processing if no data remains after filtering

    # Filter smoothing weight
    filtered_df = df.copy()
    i = 0

    while i < len(df) - 1:
        val_ini = filtered_df["Ptot"].iloc[i]
        j = i + 1

        while j < len(df) and np.abs(df["Ptot"].iloc[j] - val_ini) < seuil_poids:
            j += 1

        if j <= len(df):
            filtered_df.loc[i : j - 1, "Ptot"] = val_ini
        i = j

    df = filtered_df  # Update df with the noise-filtered data

    # Calcul du temps d'activité et du nombre de bouchée
    activity_time = 0
    peaks_x = []
    peaks_y = []

    # Peak detection
    peaks, _ = find_peaks(
        df["Ptot"], height=100, distance=5
    )  # Reduced distance for more sensitivity
    prominences = peak_prominences(df["Ptot"], peaks)[0]

    # Filter peaks based on prominence
    poids_seuil = 10  # Lowered threshold for more sensitivity
    significant_peaks_indices = np.where(prominences > poids_seuil)[0]
    significant_peaks = peaks[significant_peaks_indices]
    significant_prominences = prominences[significant_peaks_indices]

    # Further filter peaks to ensure sufficient weight difference
    weight_diff_threshold = 2  # Lowered threshold for more sensitivity
    valid_peaks_indices = []
    for idx in significant_peaks:
        if (
            (idx - 1 >= 0 and idx + 1 < len(df))
            and (
                df["Ptot"].iloc[idx] - df["Ptot"].iloc[idx - 1] > weight_diff_threshold
            )
            and (
                df["Ptot"].iloc[idx] - df["Ptot"].iloc[idx + 1] > weight_diff_threshold
            )
        ):
            # Check if weight changes significantly before and after the peak
            window = 5  # Define a window around the peak to check weight change
            if (
                (idx - window >= 0 and idx + window < len(df))
                and (
                    abs(df["Ptot"].iloc[idx - window] - df["Ptot"].iloc[idx])
                    > weight_diff_threshold
                )
                and (
                    abs(df["Ptot"].iloc[idx + window] - df["Ptot"].iloc[idx])
                    > weight_diff_threshold
                )
            ):
                valid_peaks_indices.append(idx)

    # Filter closely spaced peaks and merge windows
    min_diff = 50  # Minimum difference in indices between consecutive peaks

    stop_the_bite = True
    allPeaksFound = True
    add_peak_update_next = 0
    activity_time = 0
    final_peaks_indices = []
    merged_windows = []
    is_bite = []
    window_start, window_end = 0, 0
    Duree_activite_min = float("inf")
    Duree_activite_max = 0
    bouchees = 0
    i = 0
    while True:
        i += 1
        while i<len(df) and df["Ptot"].iloc[i] <= df["Ptot"].iloc[i - 1]:
            i += 1
        if i == len(df):
            break
        if final_peaks_indices:
            lastI = i == len(valid_peaks)
            if stop_the_bite:
                # increase the window to the left
                exploring_horizontal = 5
                for j in range(window_start, -1, -1):
                    if df["Ptot"].iloc[j] == df["Ptot"].iloc[j - 1]:
                        exploring_horizontal -= 1
                        if exploring_horizontal == 0:
                            break
                    else:
                        exploring_horizontal = 5
                        window_start = j - 1
            upTo = len(df) - 1 if lastI else valid_peaks[i]
            stop_the_bite = lastI or (upTo - final_peaks_indices[-1]) > min_diff
            if stop_the_bite:
                # increase the window to the right
                exploring_horizontal = 5
                for j in range(window_end, upTo):
                    if df["Ptot"].iloc[j] == df["Ptot"].iloc[j + 1]:
                        exploring_horizontal -= 1
                        if exploring_horizontal == 0:
                            break
                    else:
                        exploring_horizontal = 5
                        window_end = j + 1
                        if window_end == 997:
                            print(5)
                stop_the_bite = lastI or (upTo - window_end) > min_diff
            if not stop_the_bite:
                # cut into two bites if a long period of inactivity is found in the bite window
                last_activity = window_end
                for window_endi in range(window_end, upTo):
                    if (
                        df["Ptot"].iloc[window_endi]
                        != df["Ptot"].iloc[window_endi + 1]
                    ):
                        if (
                            df["time"].iloc[window_endi]
                            - df["time"].iloc[last_activity]
                            > min_inactivity
                        ):
                            stop_the_bite = True
                            window_end = last_activity
                            if window_end == 997:
                                print(5)
                            break
                        last_activity = window_endi + 1
            if stop_the_bite:
                # if window_start == 738:
                #     print(7)
                merged_windows.append((window_start, window_end))
                Duree_activity = (
                    df["time"].iloc[window_end] - df["time"].iloc[window_start]
                )
                activity_time += Duree_activity
                if Duree_activite_min > Duree_activity:
                    Duree_activite_min = Duree_activity
                if Duree_activite_max < Duree_activity:
                    Duree_activite_max = Duree_activity
                # check if the activity decreases the amount of food
                is_bite.append(
                    df["Ptot"].iloc[window_end] < df["Ptot"].iloc[window_start]
                )
                if is_bite[-1]:
                    bouchees += 1
                    # check if the food quantity has decreased before this bite
                    if (
                        len(merged_windows) > 1
                        and df["Ptot"].iloc[window_start]
                        != df["Ptot"].iloc[merged_windows[-2][1]]
                        and 10):
                        # go back to see where would be the missing bite
                        # last_quantity = df["Ptot"].iloc[merged_windows[-2][1]]
                        in_peak = False
                        for j in range(merged_windows[-2][1] + 1, window_start):
                            # if df["time"].iloc[j]>=390.95:
                            #     print(5)
                            if j>0 and df["Ptot"].iloc[j] > df["Ptot"].iloc[j - 1]:
                                # if j == 719 or j == 722 or j == 723 or j == 725:
                                #     print(7)
                                valid_peaks = np.insert(valid_peaks, firstPeakAfterPrevAct, j)
                                firstPeakAfterPrevAct += 1
                                valid_prominences = np.insert(
                                    valid_prominences, i - 1, 0
                                )
                                allPeaksFound = False
                                i += 1
                            # if df["Ptot"].iloc[j] > last_quantity + min_peak:
                            #     valid_peaks = np.insert(valid_peaks, i - 1, j)
                            #     valid_prominences = np.insert(
                            #         valid_prominences, i - 1, 0
                            #     )
                            #     allPeaksFound = False
                            #     i += 1
                            #     in_peak = True
                            #     exploring_horizontal = 5
                            # elif (
                            #     not in_peak
                            #     and df["Ptot"].iloc[j] < last_quantity
                            #     and j > 0
                            #     and df["Ptot"].iloc[j] == df["Ptot"].iloc[j - 1]
                            # ):
                            #     last_quantity = df["Ptot"].iloc[j]
                            # elif j>0 and df["Ptot"].iloc[j] == df["Ptot"].iloc[j - 1]:
                            #     if exploring_horizontal == 0:
                            #         last_quantity = df["Ptot"].iloc[j]
                            #     exploring_horizontal-=1
                            #     in_peak = False
                add_peak_update_next = not lastI
                firstPeakAfterPrevAct = i
            else:
                # Merge peaks if they are close
                window_end = valid_peaks[i]
                if window_end == 997:
                    print(5)
                if (
                    valid_prominences[i]
                    > valid_prominences[
                        np.where(valid_peaks == final_peaks_indices[-1])[0][0]
                    ]
                ):
                    final_peaks_indices[-1] = valid_peaks[i]
        if add_peak_update_next!=-1:
            final_peaks_indices.append(valid_peaks[i])
            window_start = valid_peaks[i]
            window_end = valid_peaks[i]
            add_peak_update_next = -1

    significant_peaks_x = df["time"].iloc[final_peaks_indices].values
    significant_peaks_y = df["Ptot"].iloc[final_peaks_indices].values

    # Calcul du temps d'activité
    # if len(significant_peaks_x) > 0:
    #     activity_start_time = significant_peaks_x[0]
    #     activity_end_time = significant_peaks_x[-1]
    #     activity_time = math.trunc(activity_end_time - activity_start_time)
    # else:
    #     activity_time = 0

    if bouchees:
        debut_time = merged_windows[0][0]
        fin_time = merged_windows[-1][1]

        poids_consome = df["Ptot"].iloc[debut_time] - df["Ptot"].iloc[fin_time]
        temps_repas = df["time"].iloc[fin_time] - df["time"].iloc[debut_time]
    else:
        poids_consome = 0
        temps_repas = 0
    ratio = activity_time / temps_repas if temps_repas > 0 else 0

    print(f"Le poids consommé pendant le repas est : {poids_consome}")
    print(f"La durée du repas est : {convert_time(temps_repas)}")
    print(f"Le temps d'activité est : {convert_time(activity_time)}")
    print(f"Le ratio d'activité est : {ratio}")
    print(f"Le nombre de bouchée pendant le repas est : {bouchees}")

    new_excel[fichier] = dict()
    new_excel[fichier]["Bouchees"] = bouchees
    new_excel[fichier]["Proportion_activite"] = round(ratio * 100, 1)
    new_excel[fichier]["Duree_activite_min"] = Duree_activite_min
    new_excel[fichier]["Duree_activite_max"] = Duree_activite_max
    new_excel[fichier]["Duree_activite_mean"] = round(activity_time / bouchees, 3)
    new_excel[fichier]["Duree_activite_Totale"] = activity_time
    new_excel[fichier]["Action"] = len(merged_windows)
    new_excel[fichier]["Poids_Conso"] = poids_consome
    new_excel[fichier]["Duree_Totale"] = temps_repas

    # Création de graphiques avec Plotly
    fig = Figure()
    fig.update_layout(
        title_text="Poids Total en fonction du temps",
        xaxis_title="Temps en secondes",
        yaxis_title="Poids en grammes",
    )
    fig.add_trace(Scatter(y=df["Ptot"], x=df["time"], mode="lines", name="Poids Total"))

    # Ajouter les fenêtres en rouge pour chaque pic
    for index, (start_idx, end_idx) in enumerate(merged_windows):
        # # Début de la fenêtre : recherche en arrière jusqu'à ce que le poids commence à augmenter
        # while (
        #     start_idx > 0
        #     and df["Ptot"].iloc[start_idx - 1] < df["Ptot"].iloc[start_idx]
        # ):
        #     start_idx -= 1

        # # Fin de la fenêtre : recherche en avant jusqu'à ce que le poids redevienne constant
        # while (
        #     end_idx < len(df) - 1
        #     and df["Ptot"].iloc[end_idx + 1] < df["Ptot"].iloc[end_idx]
        # ):
        #     end_idx += 1
        color = "Red" if is_bite[index] else "Gray"
        fig.add_shape(
            type="rect",
            x0=df["time"].iloc[start_idx],
            y0=df["Ptot"].min(),
            x1=df["time"].iloc[end_idx],
            y1=df["Ptot"].max(),
            line=dict(color=color, width=2),
            fillcolor=color,
            opacity=0.2,
        )

    fig.add_trace(
        Scatter(
            y=significant_peaks_y,
            x=significant_peaks_x,
            mode="markers",
            name="Pics détectés",
            marker=dict(color="red", size=8),
        )
    )
    fig.add_annotation(
        x=df["time"].iloc[-1],
        y=df["Ptot"].max(),
        text=f"Poids consommé: {poids_consome} g<br>Durée du repas: {convert_time(temps_repas)}<br>Temps d'activité sur l'assiette: {convert_time(activity_time)}<br>Ratio d'activité: {ratio:.2f}<br>Nombre de bouchées: {bouchees}",
        showarrow=False,
        align="left",
        xanchor="right",
        yanchor="top",
    )

    # valid_peaks_x = df["time"].iloc[valid_peaks].values
    # valid_peaks_y = df["Ptot"].iloc[valid_peaks].values
    # fig.add_trace(
    #     Scatter(
    #         y=valid_peaks_y,
    #         x=valid_peaks_x,
    #         mode="markers",
    #         name="Pics significatifs",
    #         marker=dict(color="blue", size=8),
    #     )
    # )

    # Enregistrement des graphiques
    filepath = os.path.join(
        dossier_graphique,
        "Graph_{}.html".format(os.path.basename(fichier).split(".")[0]),
    )
    fig.write_html(filepath)

    # Extract features for each bite and store in a list
    feature_list = []
    for start_idx, end_idx in merged_windows:
        segment = df.iloc[start_idx : end_idx + 1]
        features = extract_features(segment)
        features["BiteID"] = len(feature_list) + 1  # Unique identifier for each bite
        # Add the known label here if available, e.g., features['Label'] = 'Fork'
        feature_list.append(features)

    # Convert to DataFrame
    features_df = pd.DataFrame(feature_list)

    # Save to CSV
    features_df.to_csv("bite_features.csv", index=False)


# Open an existing workbook
workbook = openpyxl.load_workbook(excel_all_path)

# Select the active worksheet (you can also select a specific sheet by name)
sheet = workbook.active


for fichier in fichiers:
    # Iterate through each row in the column
    search_string = fichier.rsplit("\\", 1)[1].replace(".xlsx", date_folder)
    for row_num in range(1, sheet.max_row + 1):
        cell_value = sheet[f"T{row_num}"].value
        if cell_value == search_string:
            for index, key in enumerate(excel_titles, start=11):
                cell = sheet.cell(
                    row=row_num, column=index
                )  # Column 'S' is the 19th column
                cell.value = new_excel[fichier][key]
            break

# Save the changes to the workbook
workbook.save(excel_all_path)
