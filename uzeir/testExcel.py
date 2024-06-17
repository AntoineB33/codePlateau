from openpyxl import load_workbook

excel_path = r".\Resultats exp bag_couverts\Resultats exp bag_couverts\test.xlsx"

# Open an existing workbook
workbook = load_workbook(excel_path)

# Select the active worksheet (you can also select a specific sheet by name)
sheet = workbook.worksheets[0]


cell = sheet.cell(
    row=1, column=1
)
cell.value = "Repas"

# for sheet in workbook:
#   sheet._legacy_drawing = None
    

# Save the changes to the workbook
workbook.save(excel_path)
workbook.close()